from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
from pytest import mark
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants

#prefix => ön ek test_
#postfix => 
class Test_DemoClass:
    #her testten önce çağrılır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        #günün tarihini al bu tarih ilebir klasör var mı kontrol et yoksa oluştur
        self.folderPath = str(date.today())
        Path(self.folderPath).mkdir(exist_ok=True)

    #her testten sonra çağrılır
    def teardown_method(self):
        self.driver.quit()

    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook(globalConstants.dataPath)
        selectedSheet = excelFile["Sheet1"]
        totalRows = selectedSheet.max_row
        data = []

        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)

        return data
    #@mark.skip()
    @mark.parametrize(globalConstants.tupleUsrPwd,getData())
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,globalConstants.username))
        userNameInput = self.driver.find_element(By.ID,globalConstants.username)
        passwordInput = self.driver.find_element(By.ID,globalConstants.password)
        userNameInput.send_keys(username)
        passwordInput.send_keys(password)
        loginBtn = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginBtn.click()   
        errorMessage = self.driver.find_element(By.XPATH,"/html/body/div/div/div[2]/div[1]/div/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")
        assert errorMessage.text == globalConstants.incorrectLoginMsg

    @mark.parametrize(globalConstants.tupleUsrPwd,[("standard_user","secret_sauce"),  ("problem_user","secret_sauce")])
    def test_valid_login(self,username,password):
        self.driver.get(globalConstants.URL)
        self.waitForElementVisible((By.ID,globalConstants.username))
        usernameInput = self.driver.find_element(By.ID,globalConstants.username)
        passwordInput = self.driver.find_element(By.ID,globalConstants.password)
        # Action Chains
        actions = ActionChains(self.driver)
        actions.send_keys_to_element(usernameInput,username)
        actions.send_keys_to_element(passwordInput,password)
        actions.perform()
        loginBtn = self.driver.find_element(By.ID,globalConstants.loginButton)
        loginBtn.click()
        self.driver.save_screenshot(f"{self.folderPath}/test-valid-login-{username}-{password}.png")

    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))  